from openpyxl import load_workbook
wb= load_workbook('test.xlsx')
#Instructions
#This script will move all compiled data and inputed rates into recharges
#In order to run this script, you must first have a worksheet with the compiled data.
#Export from the data base the CHW and KWH data (be sure to have column names)
#Fill in data from reads for Water and Gas
#Correct any anomalies in the compiled data.

# C(july)-N(june) Use input month # to determine which column to input data.
while True:
	month=input("Input Month #(1-12): ")
	if month > 12 or month<1:
		error= 'Invalid Month, please input a month from 1-12'
		str(error)
	elif month >= 7:
		inputColumn= 3+(month-7)
		break
	else:
		inputColumn= 3+month+5
		break
# #KWH Rate
# elCap=input("Input El Cap KWH: ")
# elCapPrice=input("Input El Cap $: ")
# wilson=input("Input Wilson KWH: ")
# wilsonPrice=input("Input Wilson $: ")
# sunpower=input("Input Sunpower KWH: ")
# sunpowerPrice=input("Input Sun $: ")
# UCOP=input("input UCOP $: ")
# kwhRate= (elCap+wilson+sunpower)/(elCapPrice+wilsonPrice+UCOP)

# #Water Rate
# h2oRate= input("Input Water rate: ")

# #Gas Rate
# gasRate= input("Input Gas rate: ")


#Inputing Utility_Summary Code, three sheets to edit
#Gall R&W Utility Summary, Facilities ... ..., Library ... ...
ws= wb["Gall R&W Utility Summary"]
name=input("Input Gas Workbook name enclosed in single quotes: ")
wb2= load_workbook(name)
print(wb2.sheetnames)
ws2 = wb2["Gas Reads 2017"]
gasInput=ws2['Q9']
ws.cell(row=12, column=inputColumn).value= gasInput.value



#TODO: Find out way to compile all usage data. Maybe CSV?
#Hard to keep track if we get data from database, bills, etc.
#Example code of inputing into specific cell  d = ws.cell(row=4, column=2, value=10)
#Ask Gabriel to format Gas reads to be like Utility Recharges (Months in same columns)

#Save Changes to chosen excel sheet
wb.save('test.xlsx')